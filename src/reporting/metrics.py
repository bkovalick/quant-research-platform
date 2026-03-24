from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from io import BytesIO

from domain.portfolio.portfolio import Portfolio
from models.backtest_result import BacktestResult
from models.experiment import Experiment
from models.rebalance_problem import RebalanceProblem
from models.market_config import MarketStoreConfig, MarketStateConfig

# def deserialize_series(data) -> pd.Series:
#     """Deserialize a series from JSON round-trip (handles {index, values} format and plain dicts)."""
#     if isinstance(data, pd.Series):
#         return data
#     if isinstance(data, dict) and "index" in data and "values" in data:
#         return pd.Series(data["values"], index=data["index"])
#     if isinstance(data, dict):
#         return pd.Series(data)
#     return pd.Series(data)


# def deserialize_dataframe(data) -> pd.DataFrame:
#     """Deserialize a DataFrame from JSON round-trip."""
#     if isinstance(data, pd.DataFrame):
#         return data.copy()
#     if isinstance(data, dict) and "index" in data and "columns" in data and "values" in data:
#         return pd.DataFrame(data["values"], index=data["index"], columns=data["columns"])
#     if isinstance(data, dict):
#         try:
#             return pd.DataFrame.from_dict(data, orient="index")
#         except Exception:
#             return pd.DataFrame(data)
#     return pd.DataFrame(data)


# class ExcelGenerator:
#     def __init__(self, experiment: Experiment, buffer: BytesIO):
#         self.experiment = experiment
#         self.config = experiment.market_config
#         self.buffer = buffer

#     def generate_report(self):
#         results = self.aggregate_performance_metrics()
#         wb = Workbook()
#         default_sheet = wb.active
#         wb.remove(default_sheet)

#         if "summary" in results and results["summary"] is not None:
#             summary_ws = wb.create_sheet(title="Summary")
#             for r_idx, row in enumerate(dataframe_to_rows(results["summary"], header=True, index=False), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     summary_ws.cell(row=r_idx, column=c_idx, value=value)

#         if "time_series" in results and results["time_series"] is not None:
#             ts_ws = wb.create_sheet(title="Time Series")
#             for r_idx, row in enumerate(dataframe_to_rows(results["time_series"], header=True, index=False), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     ts_ws.cell(row=r_idx, column=c_idx, value=value)

#         if "rolling_time_series" in results and results["rolling_time_series"] is not None:
#             ts_ws = wb.create_sheet(title="Rolling Time Series")
#             for r_idx, row in enumerate(dataframe_to_rows(results["rolling_time_series"], header=True, index=False), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     ts_ws.cell(row=r_idx, column=c_idx, value=value)

#         wb.save(self.buffer)
#         self.buffer.seek(0)

#     def aggregate_performance_metrics(self):
#         """Aggregate performance metrics from multiple strategies into summary and time series DataFrames."""
#         summary_rows = []
#         portfolio_dfs = []
#         rolling_dfs = []

#         for strategy_run in self.experiment.strategy_runs:
#             strategy_name = strategy_run.strategy_name

#             # Summary
#             row = {"strategy": strategy_name}
#             for k, v in strategy_run.result.summary.items():
#                 if isinstance(v, (pd.Series, pd.DataFrame)):
#                     continue
#                 row[k] = v
#             summary_rows.append(row)

#             # Time series
#             series = strategy_run.result.series
#             if "portfolio_weights" in series:
#                 try:
#                     weights_df = deserialize_dataframe(series["portfolio_weights"])
#                     wealth_series = deserialize_series(series["portfolio_wealth_factors"])
#                     returns_series = deserialize_series(series["portfolio_returns"])
#                     turnover_series = deserialize_series(series["portfolio_turnover"])
#                     trades_series = deserialize_series(series["portfolio_trades"])

#                     weights_df = weights_df.reset_index(drop=True)
#                     min_len = min(
#                         len(weights_df), len(wealth_series),
#                         len(returns_series), len(turnover_series), len(trades_series)
#                     )
#                     weights_df = weights_df.iloc[:min_len].copy()

#                     weights_df.insert(0, "Date", pd.to_datetime(wealth_series.index[:min_len]))
#                     weights_df.insert(1, "Strategy", strategy_name)
#                     weights_df.insert(2, "WealthFactor", wealth_series.values[:min_len])
#                     weights_df.insert(3, "PortfolioReturns", returns_series.values[:min_len])
#                     weights_df.insert(4, "PortfolioTurnover", turnover_series.values[:min_len])
#                     weights_df.insert(5, "PortfolioTrades", trades_series.values[:min_len])
#                     portfolio_dfs.append(weights_df)
#                 except Exception as e:
#                     print(f"Warning: could not build time series for {strategy_name}: {e}")

#             # Rolling time series
#             if "rolling_returns" in series:
#                 try:
#                     rr = deserialize_series(series["rolling_returns"])
#                     rolling_df = pd.DataFrame({
#                         "Date": pd.to_datetime(rr.index),
#                         "Strategy": strategy_name,
#                         "RollingReturns": rr.values,
#                         "RollingVolatility": deserialize_series(series["rolling_volatility"]).values,
#                         "RollingSharpe": deserialize_series(series["rolling_sharpe_ratio"]).values,
#                         "RollingDrawdown": deserialize_series(series["rolling_drawdown"]).values,
#                         "RollingTurnover": deserialize_series(series["rolling_turnover"]).values,
#                     })
#                     rolling_dfs.append(rolling_df)
#                 except Exception as e:
#                     print(f"Warning: could not build rolling series for {strategy_name}: {e}")

#         summary_df = pd.DataFrame(summary_rows)
#         portfolio_metrics_df = pd.concat(portfolio_dfs, axis=0, ignore_index=True) if portfolio_dfs else None
#         rolling_metrics_df = pd.concat(rolling_dfs, axis=0, ignore_index=True) if rolling_dfs else None

#         return {
#             "summary": summary_df,
#             "time_series": portfolio_metrics_df,
#             "rolling_time_series": rolling_metrics_df
#         }

class MetricsCompute:
    def __init__(self):
        pass

    def compute(self,
                rebalance_problem: RebalanceProblem,
                portfolio: Portfolio,
                market_store_config: MarketStoreConfig,
                market_state_config: MarketStateConfig,
                benchmark_index: pd.Series) -> BacktestResult:
        """ Build each piece of the backtest result """
        self.rebalance_problem = rebalance_problem
        self.market_state_config = market_state_config
        self.annual_trading_days = market_state_config.annual_trading_days
        self.market_frequency = market_state_config.market_frequency
        performance_metrics = self._calculate_performance_metrics(portfolio, market_store_config, benchmark_index)
        performance_series = self._build_performance_series(performance_metrics)
        summary = self._build_summary(performance_metrics)
        return BacktestResult(
            summary=summary,
            series=performance_series
        )

    def _calculate_performance_metrics(self,
                                       portfolio: Portfolio,
                                       market_store_config: MarketStoreConfig,
                                       benchmark_index: pd.Series) -> dict:
        """Calculate performance metrics for the portfolio."""
        risk_free_rate = market_store_config.risk_free_rate
        portfolio_weights = portfolio.weights
        portfolio_trades = portfolio.weights.diff().abs().fillna(0)
        if isinstance(portfolio_trades, pd.DataFrame):
            portfolio_trades = portfolio_trades.sum(axis=1)

        portfolio_returns = portfolio.returns
        portfolio_turnover = portfolio.turnover
        wealth_factors = (1 + portfolio_returns).cumprod()
        cumulative_returns = wealth_factors - 1
        num_periods = cumulative_returns.shape[0]
        years = num_periods / self.annual_trading_days
        annualized_return = wealth_factors.iloc[-1] ** (1 / years) - 1
        annualized_volatility = portfolio_returns.std() * np.sqrt(self.annual_trading_days)

        sharpe_ratio = (
            (annualized_return - risk_free_rate) / annualized_volatility
            if annualized_volatility != 0 else 0.0
        )

        if self.annual_trading_days > 0:
            drawdown_returns = cumulative_returns.iloc[self.annual_trading_days:]
            rolling_dd = self._calculate_rolling_drawdown(drawdown_returns, self.annual_trading_days)
            rolling_dd = align_series_to_dataframe(cumulative_returns.copy(), rolling_dd)
        else:
            drawdown_returns = cumulative_returns
            rolling_dd = self._calculate_rolling_drawdown(drawdown_returns, self.annual_trading_days)
            rolling_dd = align_series_to_dataframe(cumulative_returns.copy(), rolling_dd)

        max_drawdown = abs(self._calculate_max_drawdown(drawdown_returns))
        max_drawdown_days = self._calculate_max_drawdown_days(drawdown_returns)

        performance_metrics = {
            "portfolio_wealth_factors": wealth_factors,
            "portfolio_trades": portfolio_trades,
            "portfolio_weights": portfolio_weights,
            "portfolio_returns": portfolio_returns,
            "portfolio_turnover": portfolio_turnover,
            "cumulative_returns": cumulative_returns,
            "rolling_returns": self._calculate_rolling_returns(portfolio_returns, self.annual_trading_days),
            "rolling_volatility": self._calculate_rolling_volatility(portfolio_returns, self.annual_trading_days),
            "rolling_sharpe_ratio": self._calculate_rolling_sharpe_ratio(portfolio_returns, risk_free_rate, self.annual_trading_days),
            "rolling_drawdown": rolling_dd,
            "rolling_turnover": self._calculate_rolling_turnover(portfolio_turnover, self.annual_trading_days),
            "return": annualized_return,
            "volatility": annualized_volatility,
            "sharpe_ratio": sharpe_ratio,
            "sortino_ratio": self._calculate_sortino_ratio(annualized_return, portfolio_returns, risk_free_rate),
            "max_drawdown": max_drawdown,
            "max_drawdown_days": max_drawdown_days,
            "avg_drawdown": self._calculate_avg_drawdown(drawdown_returns),
            "turnover": portfolio_turnover.mean() * self.annual_trading_days,
            "alpha": self._calculate_alpha(portfolio_returns, self.annual_trading_days, benchmark_index),
            "calmar_ratio": annualized_return / max_drawdown if max_drawdown != 0 else 0.0,
            "tracking_error": np.sqrt(
                ((portfolio_returns - benchmark_index.pct_change().fillna(0)) ** 2).mean()
            ) * np.sqrt(self.annual_trading_days),
            "win_rate": (portfolio_returns > 0).mean(),
            "loss_rate": (portfolio_returns < 0).mean(),
            "average_win": portfolio_returns[portfolio_returns > 0].mean() if (portfolio_returns > 0).any() else 0.0,
            "average_loss": portfolio_returns[portfolio_returns < 0].mean() if (portfolio_returns < 0).any() else 0.0,
            "skewness": portfolio_returns.skew(),
            "kurtosis": portfolio_returns.kurt(),
            "var_95": portfolio_returns.quantile(0.05),
            "var_97.5": portfolio_returns.quantile(0.025),
            "var_99": portfolio_returns.quantile(0.01),
            "cvar_95": portfolio_returns[portfolio_returns < portfolio_returns.quantile(0.05)].mean() if len(portfolio_returns) > 0 else 0.0,
            "cvar_97.5": portfolio_returns[portfolio_returns < portfolio_returns.quantile(0.025)].mean() if len(portfolio_returns) > 0 else 0.0,
            "cvar_99": portfolio_returns[portfolio_returns < portfolio_returns.quantile(0.01)].mean() if len(portfolio_returns) > 0 else 0.0,
            "alpha_decay": self._calculate_alpha(
                portfolio_returns[-self.annual_trading_days:],
                self.annual_trading_days,
                benchmark_index[-self.annual_trading_days:]
            ) if len(portfolio_returns) >= self.annual_trading_days else None
        }
        return performance_metrics
    
    def _calculate_sortino_ratio(self, annualized_return: float, portfolio_returns: pd.Series, risk_free_rate: float) -> float:
        """
        Annualised Sortino ratio: excess return over rf divided by downside deviation.
        Uses only negative-return periods to compute the downside std, then annualises
        by sqrt(annual_trading_days). Returns 0 when there are fewer than 2 negative
        return periods (std is NaN or zero).
        """
        downside_std = portfolio_returns[portfolio_returns < 0].std()
        if not pd.notna(downside_std) or downside_std == 0:
            return 0.0
        return (annualized_return - risk_free_rate) / (downside_std * np.sqrt(self.annual_trading_days))
    
    def _calculate_max_drawdown(self, cumulative_returns: pd.Series):
        """Calculate maximum drawdown from cumulative returns."""
        wealth = 1 + cumulative_returns
        running_max = wealth.cummax()
        drawdown = (wealth - running_max) / running_max
        return drawdown.min()

    def _calculate_max_drawdown_days(self, cumulative_returns: pd.Series) -> int:
        """Calculate the longest drawdown duration in periods (peak to recovery or end)."""
        running_max = cumulative_returns.cummax()
        is_underwater = (cumulative_returns < running_max).astype(int)
        not_underwater_cumsum = (1 - is_underwater).cumsum()
        max_streak = is_underwater.groupby(not_underwater_cumsum).sum().max()
        return int(max_streak) if not pd.isna(max_streak) else 0

    def _calculate_avg_drawdown(self, cumulative_returns: pd.Series) -> float:
        """Mean of all individual drawdown values at each point in time."""
        wealth = 1 + cumulative_returns
        running_max = wealth.cummax()
        drawdown = (wealth - running_max) / running_max
        negative_drawdowns = drawdown[drawdown < 0]
        if negative_drawdowns.empty:
            return 0.0
        result = float(negative_drawdowns.mean())
        return result if np.isfinite(result) else 0.0

    def _calculate_rolling_drawdown(self, cumulative: pd.Series, window: int):
        """Calculate rolling drawdown series over a specified window."""
        return cumulative.rolling(window, min_periods=1).apply(self._calculate_max_drawdown, raw=False)

    def _calculate_rolling_returns(self, returns: pd.Series, window: int):
        """Calculate rolling return over a specified window."""
        rolling_return = (1 + returns).rolling(window=window).apply(np.prod, raw=True) - 1
        return rolling_return

    def _calculate_rolling_volatility(self, returns: pd.Series, window: int):
        """Calculate rolling volatility over a specified window."""
        return returns.rolling(window=window).std() * np.sqrt(window)

    def _calculate_rolling_sharpe_ratio(self, returns: pd.Series, risk_free_rate: float, window: int):
        """Calculate rolling Sharpe ratio over a specified window."""
        rf_per_period = risk_free_rate / window  # de-annualise rf to match per-period rolling_mean
        rolling_mean = returns.rolling(window=window).mean()
        rolling_std = returns.rolling(window=window).std()
        sharpe = ((rolling_mean - rf_per_period) / rolling_std) * np.sqrt(window)
        return sharpe.replace([np.inf, -np.inf], np.nan)

    def _calculate_rolling_turnover(self, turnover: pd.Series, window: int):
        """Calculate rolling turnover over a specified window."""
        return turnover.rolling(window=window).mean() * np.sqrt(window)

    def _calculate_alpha(self,
                         portfolio_returns: pd.Series,
                         annualization_factor: int,
                         benchmark_index: pd.Series):
        """Calculate alpha of the portfolio against a benchmark."""
        rule = {"d": "B", "w": "W-FRI", "m": "M"}[self.market_frequency]
        benchmark = benchmark_index.resample(rule).last()
        benchmark_returns = benchmark.pct_change().fillna(0)

        aligned = pd.concat([portfolio_returns, benchmark_returns], axis=1, join='inner')
        aligned.columns = ['portfolio', 'benchmark']

        portfolio_annualized = (1 + aligned['portfolio']).prod() ** (annualization_factor / len(aligned)) - 1
        benchmark_annualized = (1 + aligned['benchmark']).prod() ** (annualization_factor / len(aligned)) - 1
        return portfolio_annualized - benchmark_annualized

    def _build_performance_series(self, performance_metrics: dict) -> dict:
        """Create performance series dictionary for BacktestResult dataclass."""
        if performance_metrics is None:
            return {}
        return {
            "portfolio_wealth_factors": performance_metrics["portfolio_wealth_factors"],
            "portfolio_trades": performance_metrics["portfolio_trades"],
            "portfolio_weights": performance_metrics["portfolio_weights"],
            "portfolio_returns": performance_metrics["portfolio_returns"],
            "portfolio_turnover": performance_metrics["portfolio_turnover"],
            "cumulative_returns": performance_metrics["cumulative_returns"],
            "rolling_returns": performance_metrics["rolling_returns"],
            "rolling_volatility": performance_metrics["rolling_volatility"],
            "rolling_sharpe_ratio": performance_metrics["rolling_sharpe_ratio"],
            "rolling_drawdown": performance_metrics["rolling_drawdown"],
            "rolling_turnover": performance_metrics["rolling_turnover"]
        }

    def _build_summary(self, performance_metrics) -> dict:
        """Create summary dictionary for BacktestResult dataclass."""
        if performance_metrics is None:
            return {}
        return {
            "return": performance_metrics["return"],
            "volatility": performance_metrics["volatility"],
            "sharpe_ratio": performance_metrics["sharpe_ratio"],
            "sortino_ratio": performance_metrics["sortino_ratio"],
            "max_drawdown": performance_metrics["max_drawdown"],
            "avg_drawdown": performance_metrics["avg_drawdown"],
            "max_drawdown_duration": performance_metrics["max_drawdown_days"],
            "turnover": performance_metrics["turnover"],
            "alpha": performance_metrics["alpha"],
            "calmar_ratio": performance_metrics["calmar_ratio"],
            "tracking_error": performance_metrics["tracking_error"],
            "win_rate": performance_metrics["win_rate"],
            "loss_rate": performance_metrics["loss_rate"],
            "average_win": performance_metrics["average_win"],
            "average_loss": performance_metrics["average_loss"],
            "skewness": performance_metrics["skewness"],
            "kurtosis": performance_metrics["kurtosis"],
            "var_95": performance_metrics["var_95"],
            "var_97.5": performance_metrics["var_97.5"],
            "var_99": performance_metrics["var_99"],
            "cvar_95": performance_metrics["cvar_95"],
            "cvar_97.5": performance_metrics["cvar_97.5"],
            "cvar_99": performance_metrics["cvar_99"],
            "alpha_decay": performance_metrics["alpha_decay"]
        }


def align_series_to_dataframe(df: pd.DataFrame, series: pd.Series) -> pd.Series:
    n = len(df) - len(series)
    nan_part = pd.Series([np.nan] * n, index=df.index[:n])
    aligned = pd.concat([nan_part, series])
    aligned = aligned.reindex(df.index)
    return aligned
