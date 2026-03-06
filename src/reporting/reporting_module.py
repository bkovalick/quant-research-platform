from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path

from domain.portfolio.portfolio import Portfolio
from models.backtest_result import BacktestResult
from models.experiment import Experiment
from models.rebalance_problem import RebalanceProblem
from models.market_config import MarketStoreConfig, MarketStateConfig

class ExcelGenerator:
    def __init__(self, experiment: Experiment, folder_name: str):
        self.experiment = experiment
        self.config = experiment.market_config
        self.folder_name = folder_name + "/" + datetime.now().strftime('%Y-%m-%d')
        self.create_folder_path(self.folder_name)

    def create_folder_path(self, folder_name: str):
        path = Path(folder_name)
        path.mkdir(parents=True, exist_ok=True)

    def generate_report(self):
        full_filename = (
            f"{self.folder_name}/backtest_report_"
            f"{self.config.start_date}_{self.config.end_date}_"
            f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        )
        results = self.aggregate_performance_metrics()

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        if "summary" in results:
            summary_ws = wb.create_sheet(title="Summary")
            summary_rows = dataframe_to_rows(results["summary"], header=True, index=False)   
            for r_idx, row in enumerate(summary_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    summary_ws.cell(row=r_idx, column=c_idx, value=value)

        if "time_series" in results:
            ts_ws = wb.create_sheet(title="Time Series")
            ts_rows = dataframe_to_rows(results["time_series"], header=True, index=False)
            for r_idx, row in enumerate(ts_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ts_ws.cell(row=r_idx, column=c_idx, value=value)

        if "rolling_time_series" in results:
            ts_ws = wb.create_sheet(title="Rolling Time Series")
            ts_rows = dataframe_to_rows(results["rolling_time_series"], header=True, index=False)
            for r_idx, row in enumerate(ts_rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ts_ws.cell(row=r_idx, column=c_idx, value=value)

        wb.save(full_filename)
        wb.close()

    def aggregate_performance_metrics(self):
        """Aggregate performance metrics from multiple strategies into summary and time series DataFrames."""
        summary_rows = []
        portfolio_dfs = []
        rolling_dfs = []

        for strategy_run in self.experiment.strategy_runs:
            strategy_name = strategy_run.strategy_name
            row = {"strategy": strategy_name}
            for k, v in strategy_run.result.summary.items():
                if isinstance(v, (pd.Series, pd.DataFrame)):
                    continue
                row[k] = v
            summary_rows.append(row)

            if "portfolio_weights" in strategy_run.result.series:
                weights_df = pd.DataFrame(strategy_run.result.series["portfolio_weights"].values, columns=strategy_run.result.series["portfolio_weights"].columns)
                weights_df.insert(0, "Date", pd.to_datetime(strategy_run.result.series["portfolio_wealth_factors"].index))
                weights_df.insert(1, "Strategy", strategy_name)
                weights_df.insert(2, "WealthFactor", strategy_run.result.series["portfolio_wealth_factors"].values)
                weights_df.insert(3, "PortfolioReturns", strategy_run.result.series["portfolio_returns"].values)
                weights_df.insert(4, "PortfolioTurnover", strategy_run.result.series["portfolio_turnover"].values)
                portfolio_dfs.append(weights_df)

            if "rolling_returns" in strategy_run.result.series:
                rolling_df = pd.DataFrame({
                    "Date": pd.to_datetime(strategy_run.result.series["rolling_returns"].index),
                    "Strategy": strategy_name,
                    "RollingReturns": strategy_run.result.series["rolling_returns"].values,
                    "RollingVolatility": strategy_run.result.series["rolling_volatility"].values,
                    "RollingSharpe": strategy_run.result.series["rolling_sharpe_ratio"].values,
                    "RollingDrawdown": strategy_run.result.series["rolling_drawdown"].values,
                    "RollingTurnover": strategy_run.result.series["rolling_turnover"].values                    
                })
                rolling_dfs.append(rolling_df)

        summary_df = pd.DataFrame(summary_rows)
        if portfolio_dfs:
            portfolio_metrics_df = pd.concat(portfolio_dfs, axis=0, ignore_index=True)
        else:
            portfolio_metrics_df = None

        if rolling_dfs:
            rolling_metrics_df = pd.concat(rolling_dfs, axis=0, ignore_index=True)
        else:
            rolling_metrics_df = None

        return {
            "summary": summary_df,
            "time_series": portfolio_metrics_df, 
            "rolling_time_series": rolling_metrics_df
        }

class MetricsCompute:
    def __init__(self):
        pass

    def compute(self, 
                rebalance_problem: RebalanceProblem, 
                portfolio: Portfolio, 
                market_store_config: MarketStoreConfig,
                market_state_config: MarketStateConfig,
                benchmark_index: pd.Series) -> BacktestResult:
        """ Build each piece of the backtest result """
        self.rebalance_problem = rebalance_problem
        self.market_state_config = market_state_config
        self.annual_trading_days = market_state_config.annual_trading_days
        self.market_frequency = market_state_config.market_frequency
        performance_metrics = self._calculate_performance_metrics(portfolio, market_store_config, benchmark_index)
        performance_series = self._build_performance_series(performance_metrics)
        summary = self._build_summary(performance_metrics)
        return BacktestResult(
            summary=summary,
            series=performance_series
        )

    def _calculate_performance_metrics(self, 
                                       portfolio: Portfolio, 
                                       market_str_cfg: MarketStoreConfig,
                                       benchmark_index: pd.Series) -> dict:
        """Calculate performance metrics for the portfolio."""
        portfolio_weights = portfolio.weights
        portfolio_trades = portfolio.weights.diff().fillna(0)
        portfolio_returns = portfolio.returns
        portfolio_turnover = portfolio.turnover
        wealth_factors = (1 + portfolio_returns).cumprod()
        cumulative_returns = wealth_factors - 1
        num_periods = cumulative_returns.shape[0]
        years = num_periods / self.annual_trading_days
        annualized_return = wealth_factors.iloc[-1] ** (1 / years) - 1
        annualized_volatility = portfolio_returns.std() * np.sqrt(self.annual_trading_days)

        sharpe_ratio = (
            annualized_return / annualized_volatility 
            if annualized_volatility != 0 else 0.0
        )

        if self.annual_trading_days > 0:
            drawdown_returns = cumulative_returns.iloc[self.annual_trading_days:]
            rolling_dd = self._calculate_rolling_drawdown(drawdown_returns, self.annual_trading_days)
            rolling_dd = align_series_to_dataframe(cumulative_returns.copy(), rolling_dd)
        else:
            drawdown_returns = cumulative_returns
            rolling_dd = self._calculate_rolling_drawdown(drawdown_returns, self.annual_trading_days)
            rolling_dd = align_series_to_dataframe(cumulative_returns.copy(), rolling_dd)
        max_drawdown = abs(self._calculate_max_drawdown(drawdown_returns))

        performance_metrics = {
            "portfolio_wealth_factors": wealth_factors,
            "portfolio_trades": portfolio_trades,
            "portfolio_weights": portfolio_weights,
            "portfolio_returns": portfolio_returns,
            "portfolio_turnover": portfolio_turnover,
            "cumulative_returns": cumulative_returns,            
            "rolling_returns": self._calculate_rolling_returns(
                portfolio_returns, self.annual_trading_days),
            "rolling_volatility": self._calculate_rolling_volatility(
                portfolio_returns, self.annual_trading_days),
            "rolling_sharpe_ratio": self._calculate_rolling_sharpe_ratio(
                portfolio_returns, self.annual_trading_days),
            "rolling_drawdown": rolling_dd,
            "rolling_turnover": self._calculate_rolling_turnover(
                portfolio_turnover, self.annual_trading_days),
            "return": annualized_return,
            "volatility": annualized_volatility,
            "sharpe_ratio": sharpe_ratio,
            "sortino_ratio": (annualized_return / portfolio_returns[portfolio_returns < 0].std() * \
                np.sqrt(self.annual_trading_days)) if portfolio_returns[portfolio_returns < 0].std() != 0 else 0,
            "max_drawdown": max_drawdown,
            "turnover": portfolio_turnover.mean() * self.annual_trading_days,
            "alpha": self._calculate_alpha(portfolio_returns, self.annual_trading_days, market_str_cfg, benchmark_index),
            "calmar_ratio": annualized_return / max_drawdown if max_drawdown != 0 else 0.0,
            "tracking_error": np.sqrt(((portfolio_returns - benchmark_index.pct_change().fillna(0)) ** 2).mean()) * \
                np.sqrt(self.annual_trading_days),
            "win_rate": (portfolio_returns > 0).mean(),
            "loss_rate": (portfolio_returns < 0).mean(),
            "average_win": portfolio_returns[portfolio_returns > 0].mean() if (portfolio_returns > 0).any() else 0.0,
            "average_loss": portfolio_returns[portfolio_returns < 0].mean() if (portfolio_returns < 0).any() else 0.0,
            "value_at_risk_95": portfolio_returns.quantile(0.05),
            "value_at_risk_97.5": portfolio_returns.quantile(0.025),
            "value_at_risk_99": portfolio_returns.quantile(0.01),
            "conditional_value_at_risk_95": portfolio_returns[portfolio_returns < \
                portfolio_returns.quantile(0.05)].mean() if len(portfolio_returns) > 0 else 0.0,
            "conditional_value_at_risk_97.5": portfolio_returns[portfolio_returns < \
                portfolio_returns.quantile(0.025)].mean() if len(portfolio_returns) > 0 else 0.0,
            "conditional_value_at_risk_99": portfolio_returns[portfolio_returns < \
                portfolio_returns.quantile(0.01)].mean() if len(portfolio_returns) > 0 else 0.0,
            "alpha_decay": self._calculate_alpha(portfolio_returns[-self.annual_trading_days:], \
                self.annual_trading_days, market_str_cfg, benchmark_index[-self.annual_trading_days:]) \
                    if len(portfolio_returns) >= self.annual_trading_days else None
        }
        return performance_metrics

    def _calculate_max_drawdown(self, cumulative_returns: pd.Series):
        """Calculate maximum drawdown from cumulative returns."""
        running_max = cumulative_returns.cummax()
        drawdown = (cumulative_returns - running_max) / running_max
        return drawdown.min()

    def _calculate_rolling_drawdown(self, cumulative: pd.Series, window: int):
        """Calculate rolling drawdown series over a specified window."""
        return cumulative.rolling(window, min_periods=1).apply(self._calculate_max_drawdown, raw=False)

    def _calculate_rolling_returns(self, returns: pd.Series, window: int):
        """Calculate rolling return over a specified window."""
        annualization_factor = window
        rolling_return = (1 + returns).rolling(window=window).apply(np.prod, raw=True) - 1
        rolling_return = (1 + rolling_return) ** (annualization_factor / window) - 1
        return rolling_return

    def _calculate_rolling_volatility(self, returns: pd.Series, window: int):
        """Calculate rolling volatility over a specified window."""
        annualization_factor = window
        rolling_std = returns.rolling(window=window).std()
        rolling_volatility = rolling_std * np.sqrt(annualization_factor)
        return rolling_volatility

    def _calculate_rolling_sharpe_ratio(self, returns: pd.Series, window: int):
        """Calculate rolling Sharpe ratio over a specified window."""
        annualization_factor = window
        rolling_mean = returns.rolling(window=window).mean()
        rolling_std = returns.rolling(window=window).std()
        rolling_sharpe = (rolling_mean / rolling_std) * np.sqrt(annualization_factor)
        return rolling_sharpe

    def _calculate_rolling_turnover(self, turnover: pd.Series, window: int):
        """Calculate rolling turnover over a specified window."""
        annualization_factor = window
        return turnover.rolling(window=window).mean() * np.sqrt(annualization_factor)

    def _calculate_alpha(self, 
                         portfolio_returns: pd.Series, 
                         annualization_factor: int,
                         market_str_cfg: MarketStoreConfig,
                         benchmark_index: pd.Series):
        """Calculate alpha of the portfolio against a benchmark (S&P 500)."""
        rule = {"w": "W-FRI", "m": "M"}[self.market_frequency]
        benchmark = benchmark_index.resample(rule).last()
        benchmark_returns = benchmark.pct_change().fillna(0)

        if(len(benchmark_returns) != len(portfolio_returns)):
            return

        aligned = pd.concat([portfolio_returns, benchmark_returns], axis=1, join='inner')
        aligned.columns = ['portfolio', 'benchmark']

        portfolio_annualized = (1 + aligned['portfolio']).prod() ** (annualization_factor/len(aligned)) - 1
        benchmark_annualized = (1 + aligned['benchmark']).prod() ** (annualization_factor/len(aligned)) - 1
        alpha = portfolio_annualized - benchmark_annualized
        return alpha

    def _build_performance_series(self, performance_metrics: dict) -> dict:
        """Create performance series dictionary for BacktestResult dataclass"""
        if performance_metrics is None:
            return
                
        performance_series = {
            "portfolio_wealth_factors": performance_metrics["portfolio_wealth_factors"],
            "portfolio_trades": performance_metrics["portfolio_trades"],
            "portfolio_weights": performance_metrics["portfolio_weights"],
            "portfolio_returns": performance_metrics["portfolio_returns"],
            "portfolio_turnover": performance_metrics["portfolio_turnover"],
            "cumulative_returns": performance_metrics["cumulative_returns"],
            "rolling_returns": performance_metrics["rolling_returns"],
            "rolling_volatility": performance_metrics["rolling_volatility"],
            "rolling_sharpe_ratio": performance_metrics["rolling_sharpe_ratio"],
            "rolling_drawdown": performance_metrics["rolling_drawdown"],
            "rolling_turnover": performance_metrics["rolling_turnover"]
        }
        return performance_series

    def _build_summary(self, performance_metrics) -> dict:
        """Create summary dictionary for BacktestResult dataclass"""
        if performance_metrics is None:
            return
        
        performance_summary = {
            "return": performance_metrics["return"],
            "volatility": performance_metrics["volatility"],
            "sharpe_ratio": performance_metrics["sharpe_ratio"],
            "sortino_ratio": performance_metrics["sortino_ratio"],
            "max_drawdown": performance_metrics["max_drawdown"],
            "turnover": performance_metrics["turnover"],
            "alpha": performance_metrics["alpha"],
            "calmar_ratio": performance_metrics["calmar_ratio"],
            "tracking_error": performance_metrics["tracking_error"],
            "win_rate": performance_metrics["win_rate"],
            "loss_rate": performance_metrics["loss_rate"],
            "average_win": performance_metrics["average_win"],
            "average_loss": performance_metrics["average_loss"],
            "value_at_risk_95": performance_metrics["value_at_risk_95"],
            "value_at_risk_97.5": performance_metrics["value_at_risk_97.5"],
            "value_at_risk_99": performance_metrics["value_at_risk_99"],
            "conditional_value_at_risk_95": performance_metrics["conditional_value_at_risk_95"],
            "conditional_value_at_risk_97.5": performance_metrics["conditional_value_at_risk_97.5"],
            "conditional_value_at_risk_99": performance_metrics["conditional_value_at_risk_99"],
            "alpha_decay": performance_metrics["alpha_decay"]
        }
        return performance_summary

def align_series_to_dataframe(df: pd.DataFrame, series: pd.Series) -> pd.Series:
    n = len(df) - len(series)
    nan_part = pd.Series([np.nan]*n, index=df.index[:n])
    aligned = pd.concat([nan_part, series])
    aligned = aligned.reindex(df.index)
    return aligned